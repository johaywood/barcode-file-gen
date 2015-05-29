from hubarcode.datamatrix import DataMatrixEncoder
import sys, os
from openpyxl import load_workbook
from PIL import Image

subjects = []
study_num = sys.argv[1]
subj_list_file = sys.argv[2]

class Subject:

  def __init__(self, subj_id, subj_num, barcode):
    self.subj_id = subj_id
    self.subj_num = subj_num
    self.barcode = barcode
    self.barcode_color = self.get_barcode_color(barcode)
    self.barcode_path = self.get_barcode_path(barcode)
    self.barcode_img = self.get_barcode_image(barcode)
    self.add_colored_border(self.barcode_path)
    subjects.append(self)

  def get_barcode_color(self, barcode):
    color_rgb = ''
    let = str(self.barcode[:2])
    num = int(self.barcode[-4:])
    if let < 'FF':
      color_rgb = (255, 255, 0)
    elif let == 'FF' and num < 2000:
      color_rgb = (255, 255, 0)
    elif let < 'KK':
      color_rgb = (255, 229, 204)
    elif let == 'KK' and num < 4000:
      color_rgb = (255, 229, 204)
    elif let < 'PP':
      color_rgb = (255, 128, 0)
    elif let == 'PP' and num < 6000:
      color_rgb = (255, 128, 0)
    elif let < 'UU':
      color_rgb = (0, 0, 255)
    elif let == 'UU' and num < 8000:
      color_rgb = (0, 0, 255)
    elif let < 'ZZ':
      color_rgb = (0, 225, 0)
    elif let == 'ZZ' and num < 10000:
      color_rgb = (0, 225, 0)
    return color_rgb

  def get_barcode_path(self, barcode):
    cwd = os.getcwd()
    path = cwd + '\\' + study_num + '\\' + barcode + '.png'
    return path

  def get_barcode_image(self, barcode):
    barcode_img = DataMatrixEncoder(barcode)
    barcode_img.save(self.barcode_path)
    return barcode_img

  def add_colored_border(self, barcode):
    img = Image.open(barcode)
    img = img.convert('RGB')
    px = img.load()
    for x in range(img.size[0]):
      for y in range(img.size[1]):
        if px[x, y] == (150, 150, 150):
          px[x, y] = self.barcode_color
    img.save(self.barcode_path)

def add_subjects(subj_list):
  ensure_dir('%s/' % study_num)
  wb = load_workbook(filename = subj_list)
  ws = wb.active
  num_subjects = ws.max_row
  for x in range(0, num_subjects - 1):
    Subject(ws.cell('A%s' % (x + 2)).value, ws.cell('B%s' % (x + 2)).value, ws.cell('C%s' % (x + 2)).value)
    ws.cell('D%s' % (x + 2)).value = subjects[x].barcode_path
  wb.save(subj_list)

def ensure_dir(f):
    d = os.path.dirname(f)
    if not os.path.exists(d):
        os.makedirs(d)

add_subjects(subj_list_file)
    
